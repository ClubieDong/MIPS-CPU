import openpyxl as xl
import os
import re
import pyperclip as clip


def GetPipelineSignals(sheet, col):
    signals = []
    i = 2
    while True:
        name = sheet.cell(i, col).value
        width = sheet.cell(i, col + 1).value
        if name is None or width is None:
            break
        signals.append((name, width))
        i += 1
    return signals


def FormatWidth(width):
    if width == 1:
        return "      "
    return f"[{width - 1:>2}:0]"


def GeneratePipelineReg(stage1, stage2, signals):
    lines = []
    lines.append(f"// This file is automatically generated by Script.py.")
    lines.append(f"")
    lines.append(f"module {stage1}_{stage2}(")
    lines.append(f"    input         clk,")
    lines.append(f"    input         rst,")
    lines.append(f"    input         flush,")
    lines.append(f"    input         stall,")
    for i in signals:
        lines.append(f"    input  {FormatWidth(i[1])} {stage1}_{i[0]},")
    for i in signals:
        lines.append(f"    output {FormatWidth(i[1])} {stage2}_{i[0]},")
    lines[-1] = lines[-1][:-1]
    lines.append(f");")
    for i in signals:
        lines.append(f"    reg    {FormatWidth(i[1])} {i[0]}Reg;")
    lines.append(f"")
    lines.append(f"    always @ (posedge clk)")
    lines.append(f"    begin")
    lines.append(f"        if (rst)")
    lines.append(f"            pc4Reg <= 0;")
    lines.append(f"        if (rst || flush)")
    lines.append(f"        begin")
    for i in signals:
        if i[0] != "pc4":
            lines.append(f"            {i[0]}Reg <= 0;")
    lines.append(f"        end")
    lines.append(f"        else if (!stall)")
    lines.append(f"        begin")
    for i in signals:
        lines.append(f"            {i[0]}Reg <= {stage1}_{i[0]};")
    lines.append(f"        end")
    lines.append(f"    end")
    lines.append(f"")
    for i in signals:
        lines.append(f"    assign {stage2}_{i[0]} = {i[0]}Reg;")
    lines.append(f"")
    lines.append(f"endmodule")
    for i in range(len(lines)):
        lines[i] += "\n"
    with open(f"Source/Registers/PipelineRegisters/{stage1}_{stage2}.v", "w") as f:
        f.writelines(lines)


def GenerateControl(sheet):
    signals = []
    i = 8
    while True:
        name = sheet.cell(3, i).value
        width = sheet.cell(4, i).value
        if name is None or width is None:
            break
        signals.append((name, width))
        i += 1
    totalLen = 0
    for i in signals:
        totalLen += i[1]
    lines = []
    lines.append(f"// This file is automatically generated by Script.py.")
    lines.append(f"")
    lines.append(f"module Control(")
    lines.append(f"    input  [31:0] instr,")
    for i in signals:
        lines.append(f"    output {FormatWidth(i[1])} {i[0]},")
    lines[-1] = lines[-1][:-1]
    lines.append(f");")
    names = ", ".join([i[0] for i in signals])
    lines.append(f"    assign {{{names}}} =")
    i = 5
    while True:
        name = sheet.cell(i, 1).value
        if name is None:
            break
        line = "        "
        value = sheet.cell(i, 2).value
        line += f"instr[31:26] == 6'b{value:06}" if isinstance(value, int) else " " * 25
        value = sheet.cell(i, 3).value
        line += f" && instr[25:21] == 5'b{value:05}" if isinstance(value, int) else " " * 28
        value = sheet.cell(i, 4).value
        line += f" && instr[20:16] == 5'b{value:05}" if isinstance(value, int) else " " * 28
        value = sheet.cell(i, 5).value
        line += f" && instr[15:11] == 5'b{value:05}" if isinstance(value, int) else " " * 28
        value = sheet.cell(i, 6).value
        line += f" && instr[10:6] == 5'b{value:05}" if isinstance(value, int) else " " * 27
        value = sheet.cell(i, 7).value
        line += f" && instr[5:0] == 6'b{value:06}" if isinstance(value, int) else " " * 27
        line += f" ? {totalLen}'b"
        for j in range(len(signals)):
            value = sheet.cell(i, j + 8).value
            line += f"{value:0{signals[j][1]}}_" if isinstance(value, int) else "X" * signals[j][1] + "_"
        line = line[:-1] + f" : // {name}"
        lines.append(line)
        i += 1
    lines[-1] = lines[-1].replace("?", " ").replace(":", ";")
    lines.append("")
    lines.append("endmodule")
    for i in range(len(lines)):
        lines[i] += "\n"
    with open(f"Source/Control.v", "w") as f:
        f.writelines(lines)

def Instantiation(file):
    text = file.read()
    modulePattern = re.compile(r"module (\w*)\(")
    paraPattern = re.compile(r"(input |output) (      |\[[\d| ]\d:0\]) (\w*)(,?)( //.*)?")
    name = modulePattern.findall(text)[0]
    paras = paraPattern.findall(text)
    maxLen = 0
    for i in paras:
        if len(i[2]) > maxLen:
            maxLen = len(i[2])
    lines = []
    lines.append(f"    {name} {name}")
    lines.append(f"    (")
    for i in paras:
        lines.append(f"        .{i[2]:<{maxLen}} (XXXXX){i[3]:1} // {i[0]} {i[1]}")
    lines.append(f"    );")
    lines.append(f"")
    for i in range(len(lines)):
        lines[i] += "\n"
    return lines


excel = xl.load_workbook("真值表.xlsx")

sheet_piplineRegs = excel["PipelineRegs"]
GeneratePipelineReg("IF", "ID", GetPipelineSignals(sheet_piplineRegs, 1))
GeneratePipelineReg("ID", "EX", GetPipelineSignals(sheet_piplineRegs, 4))
GeneratePipelineReg("EX", "MEM", GetPipelineSignals(sheet_piplineRegs, 7))
GeneratePipelineReg("MEM", "WB", GetPipelineSignals(sheet_piplineRegs, 10))

sheet_control = excel["Control"]
GenerateControl(sheet_control)

lines = []
with open("Source/Control/PipelineControl.v", "r") as file:
    lines += Instantiation(file)
# IF
with open("Source/PC.v", "r") as file:
    lines += Instantiation(file)
with open("Source/Memory/InstructionMemory.v", "r") as file:
    lines += Instantiation(file)
# ID
with open("Source/Registers/PipelineRegisters/IF_ID.v", "r") as file:
    lines += Instantiation(file)
with open("Source/Control/Control.v", "r") as file:
    lines += Instantiation(file)
with open("Source/Registers/RegFile.v", "r") as file:
    lines += Instantiation(file)
with open("Source/Branch.v", "r") as file:
    lines += Instantiation(file)
with open("Source/Extend.v", "r") as file:
    lines += Instantiation(file)
# EXE
with open("Source/Registers/PipelineRegisters/ID_EX.v", "r") as file:
    lines += Instantiation(file)
with open("Source/ALU.v", "r") as file:
    lines += Instantiation(file)
with open("Source/MulDiv.v", "r") as file:
    lines += Instantiation(file)
# MEM
with open("Source/Registers/PipelineRegisters/EX_MEM.v", "r") as file:
    lines += Instantiation(file)
with open("Source/Memory/DataMemory.v", "r") as file:
    lines += Instantiation(file)
with open("Source/CP0.v", "r") as file:
    lines += Instantiation(file)
with open("Source/Registers/HiLo.v", "r") as file:
    lines += Instantiation(file)
# WB
with open("Source/Registers/PipelineRegisters/MEM_WB.v", "r") as file:
    lines += Instantiation(file)
clip.copy("".join(lines))
print("Generated instantiation code has been copied into clipboard.")